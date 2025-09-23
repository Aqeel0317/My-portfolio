from flask import Flask, render_template, request, redirect, url_for
from flask_wtf import FlaskForm
from flask_wtf.csrf import CSRFProtect
from wtforms import StringField, IntegerField, SubmitField, FieldList, FormField, TextAreaField, DecimalField, validators
import webbrowser
from threading import Timer
from decimal import Decimal
from openpyxl.utils.datetime import from_excel
import openpyxl
import pdfkit
import os
import sys
import datetime
import logging
from openpyxl.styles import NamedStyle

app = Flask(__name__)
app.secret_key = 'super-secret-key'
app.config['WTF_CSRF_ENABLED'] = True

# Set up logging
logging.basicConfig(level=logging.INFO)

def resource_path(relative_path):
    """ Get absolute path to resource, works for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

# Update the path to wkhtmltopdf
config = pdfkit.configuration(wkhtmltopdf=resource_path('wkhtmltopdf/bin/wkhtmltopdf.exe'))

csrf = CSRFProtect(app)

# Define the Excel file path
EXCEL_FILE = os.path.join(os.getcwd(), 'invoices.xlsx')

# Ensure the Excel file exists
def ensure_excel_file():
    logging.info(f"Ensuring Excel file at: {EXCEL_FILE}")
    if not os.path.exists(EXCEL_FILE):
        logging.info("Excel file does not exist. Creating a new one.")
        wb = openpyxl.Workbook()
        ws_student = wb.create_sheet("students")
        ws_invoice = wb.create_sheet("Invoices")
        ws_details = wb.create_sheet("InvoiceDetails")
        
        # Add headers for each sheet
        ws_student.append(['ID', 'Name', 'Address', 'Phone','Date',])
        ws_invoice.append(['ID', 'roll_no', 'total',])
        ws_details.append(['ID', 'InvoiceID', 'roll_no', 'name', 'subject', 'total', 'obtained'])
        
        try:
            wb.save(EXCEL_FILE)
            logging.info("Excel file created successfully.")
        except Exception as e:
            logging.error(f"Error saving Excel file: {e}")
    else:
        logging.info("Excel file already exists.")

ensure_excel_file()

# WTForms for input validation
class ProductForm(FlaskForm):
    qty = IntegerField('roll_no', validators=[validators.Optional()])
    brand = StringField('name', validators=[validators.Optional()])
    model_item = StringField('subject', validators=[validators.Optional()])
    details = TextAreaField('total', validators=[validators.Optional()])
    amount = DecimalField('obtained', validators=[validators.DataRequired(), validators.NumberRange(min=0)], places=2)

class InvoiceForm(FlaskForm):
    name = StringField('Name', validators=[validators.Optional()])
    address = StringField('Address', validators=[validators.Optional()])
    phone = StringField('Phone')
    submit = SubmitField('Save and Generate Invoice')

# Utility function to get new ID for Excel rows
def get_new_id(sheet):
    return sheet.max_row

@app.route('/', methods=['GET', 'POST'])
def invoice():
    form = InvoiceForm()
    if request.method == 'POST' and form.validate_on_submit():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws_student = wb['students']
        ws_invoice = wb['Invoices']
        ws_details = wb['InvoiceDetails']

        student_id = get_new_id(ws_student)

        # Create a date style for Excel
        date_style = NamedStyle(name="datetime", number_format="YYYY-MM-DD HH:MM:SS")

        # student data with proper date handling
        student_data = [
            student_id, form.name.data, form.address.data, form.phone.data, form.homephone.data,
            form.email.data, form.city.data, form.state.data, form.zipcode.data,
            form.leaseid.data, datetime.datetime.now(), form.saleperson.data, form.floormodel.data,
            form.customodr.data, form.financeco.data
        ]
        ws_student.append(student_data)

        # Apply the date style to the date cell (11th column for date)
        date_column_index = 11  # Adjust if necessary
        date_cell = ws_student.cell(row=ws_student.max_row, column=date_column_index)
        date_cell.style = date_style

        total_amount = sum(Decimal(product_data['amount']) for product_data in form.products.data)
        tax = Decimal(form.tax.data or 0.0)
        cost = Decimal(form.cost.data or 0.0)
        warranty = Decimal(form.warranty.data or 0.0)  # Handle optional fields
        delivery = Decimal(form.delivery.data or 0.0)
        setup = Decimal(form.setup.data or 0.0)  # Handle optional fields
        discount = Decimal(form.discount.data or 0.0)
        deposit = Decimal(form.deposit.data or 0.0)

        grand_total = (total_amount + tax + warranty + delivery + setup) - discount
        balance_due = grand_total - deposit

        invoice_id = get_new_id(ws_invoice)
        invoice_data = [invoice_id, student_id, total_amount, tax, cost, warranty, delivery, setup, discount, deposit, grand_total, balance_due]
        ws_invoice.append(invoice_data)

        for product_data in form.products.data:
            product_id = get_new_id(ws_details)
            detail_data = [product_id, invoice_id, product_data['qty'], product_data['brand'], product_data['model_item'], product_data['details'], product_data['amount']]
            ws_details.append(detail_data)

        try:
            wb.save(EXCEL_FILE)
            logging.info("Data saved to Excel file successfully.")
        except Exception as e:
            logging.error(f"Error saving data to Excel file: {e}")

        return redirect(url_for('view_invoice', invoice_id=invoice_id))

    return render_template('add_invoice_item.html', form=form)

@app.route('/view_invoice/<int:invoice_id>')
def view_invoice(invoice_id):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_invoice = wb['Invoices']
    ws_student = wb['students']
    ws_details = wb['InvoiceDetails']

    invoice = None
    student = None
    invoice_details = []

    # Find the invoice by ID
    for row in ws_invoice.iter_rows(values_only=True):
        if row[0] == invoice_id:
            invoice = row
            student_id = row[1]
            break

    # Find the student by ID
    for row in ws_student.iter_rows(values_only=True):
        if row[0] == student_id:
            student = list(row)  # Convert tuple to list so we can modify it
            # Convert Excel serial date to Python datetime object (assuming it's in the 10th index)
            if isinstance(student[10], (int, float)):  # Serial date is typically a float or int
                student[10] = from_excel(student[10])  # Convert serial date to Python datetime
            break

    # Find all the details for the given invoice
    for row in ws_details.iter_rows(values_only=True):
        if row[1] == invoice_id:
            invoice_details.append(row)

    # Render the invoice to HTML
    rendered_html = render_template('invoice.html', invoice=invoice, student=student, invoice_details=invoice_details)
    pdf_directory = os.path.join(os.getcwd(), 'pdfs')
    os.makedirs(pdf_directory, exist_ok=True)  # Create pdf directory if it doesn't exist
    pdf_file_path = os.path.join(pdf_directory, f'invoice_{invoice_id}.pdf')

    # Generate and save the PDF with the configuration
    options = {
        'no-stop-slow-scripts': None,
        'disable-smart-shrinking': None,
    }

    try:
        pdfkit.from_string(rendered_html, pdf_file_path, configuration=config, options=options)
        logging.info(f"PDF generated successfully: {pdf_file_path}")
    except Exception as e:
        logging.error(f"Error generating PDF: {e}")

    # Return the rendered HTML page for viewing the invoice
    return render_template('invoice.html', invoice=invoice, student=student, invoice_details=invoice_details)

    
@app.route('/delete_product/<int:product_id>/<int:invoice_id>')
def delete_product(product_id, invoice_id):
    # Load the workbook and get the sheets
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_details = wb['InvoiceDetails']

    # Find and delete the product from the InvoiceDetails
    for row in ws_details.iter_rows(min_row=2):
        if row[0].value == product_id:
            ws_details.delete_rows(row[0].row)
            break

    wb.save(EXCEL_FILE)
    return redirect(url_for('view_invoice', invoice_id=invoice_id))    


def open_browser():
    webbrowser.open_new('http://127.0.0.1:5000/')

if __name__ == '__main__':
    Timer(1, open_browser).start()
    app.run(debug=True)
