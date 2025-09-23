from flask import Flask, render_template, request, redirect, url_for, flash
import openpyxl
import os
import threading
# import webview # Commented out if not running in webview context for now
from collections import defaultdict
from werkzeug.utils import secure_filename  # For secure file names
from datetime import datetime, date
import webview
app = Flask(__name__)
app.secret_key = "secret_key"  # For flashing messages
data_dir = "data"

if not os.path.exists(data_dir):
    os.makedirs(data_dir)

# Define upload folder for pictures (stored in static/uploads)
UPLOAD_FOLDER = os.path.join(app.root_path, 'static', 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Allowed file extensions for student pictures.
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Paths for Excel files
students_file = os.path.join(data_dir, "students.xlsx")

# Initialize Excel file with updated headers including "Admission Date" and "Picture"
def initialize_excel_file(file_path, headers):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)  # Add header row
        wb.save(file_path)

# Header: "Admission Date", "Picture"
initialize_excel_file(students_file, [
    "ID", "Name", "Father Name", "Contact Number", "Grade", "Monthly Fee", "Months Paid",
    "Paid Fee", "Dues", "Paid Dates", "Annual Charge", "Paid Annual Charge", "Admission Fee",
    "Paid Admission Fee", "Admission Date", "Exam Charge", "Paid Exam Charge", "Remarks", "Picture"
])


@app.route('/')
def index():
    return render_template('index.html')


# def start_flask(): # Commented out if not running in webview context
#     app.run(host='127.0.0.1', port=5000)


@app.route('/add_student', methods=['GET', 'POST'])
def add_student():
    if request.method == 'POST':
        student_id = request.form.get('id', '').strip()
        name = request.form.get('name', '').strip()
        if not student_id or not name:
            flash("ID and Name are required.", "error")
            return redirect(url_for('add_student'))

        fathers_name = request.form.get('fathers_name', '').strip()
        contact_number = request.form.get('contact_number', '').strip()
        grade = request.form.get('grade', '').strip()
        monthly_fee = float(request.form['monthly_fee']) if request.form.get('monthly_fee', '').strip() else 0.0
        annual_charge = float(request.form['annual_charge']) if request.form.get('annual_charge', '').strip() else 0.0
        admission_fee = float(request.form['admission_fee']) if request.form.get('admission_fee', '').strip() else 0.0
        exam_charge = float(request.form['exam_charge']) if request.form.get('exam_charge', '').strip() else 0.0
        paid_annual_charge = float(request.form['paid_annual_charge']) if request.form.get('paid_annual_charge', '').strip() else 0.0
        paid_admission_fee = float(request.form['paid_admission_fee']) if request.form.get('paid_admission_fee', '').strip() else 0.0
        paid_exam_charge = float(request.form['paid_exam_charge']) if request.form.get('paid_exam_charge', '').strip() else 0.0
        paid_fee = float(request.form['paid_fee']) if request.form.get('paid_fee', '').strip() else 0.0
        
        # For 'Paid Dates', align with 'Months Paid'
        months_paid_list = request.form.getlist('months_paid')
        months_paid_str = ", ".join(months_paid_list)
        
        # If a single 'paid_date' is provided, replicate it for all months paid in this transaction
        single_paid_date_form = request.form.get('paid_date', '').strip()
        paid_dates_list = [single_paid_date_form] * len(months_paid_list)
        paid_dates_str = ", ".join(paid_dates_list)

        admission_date = request.form.get('admission_date', '').strip()

        total_months_fee = len(months_paid_list) * monthly_fee
        dues = (total_months_fee + annual_charge + admission_fee + exam_charge) - (
            paid_fee + paid_annual_charge + paid_admission_fee + paid_exam_charge)

        picture_file = request.files.get('picture')
        picture_filename = ""
        if picture_file and picture_file.filename != "":
            if allowed_file(picture_file.filename):
                filename = secure_filename(picture_file.filename)
                picture_filename = f"{student_id}_{filename}"
                picture_file.save(os.path.join(app.config['UPLOAD_FOLDER'], picture_filename))
            else:
                flash("Invalid file type for picture. Allowed types: png, jpg, jpeg, gif", "error")
                return redirect(url_for('add_student'))

        wb = openpyxl.load_workbook(students_file)
        ws = wb.active
        ws.append([
            student_id, name, fathers_name, contact_number, grade, monthly_fee, months_paid_str,
            paid_fee, dues, paid_dates_str, annual_charge, paid_annual_charge, # Use paid_dates_str
            admission_fee, paid_admission_fee, admission_date, exam_charge, paid_exam_charge, "", picture_filename
        ])
        wb.save(students_file)

        flash("Student added successfully!", "success")
        return redirect(url_for('manage_students'))

    return render_template('add_student.html')


@app.route('/students', methods=['GET', 'POST'])
def manage_students():
    search_query = request.args.get('search', '').lower().strip()
    grade_filter = request.args.get('grade', '').strip()
    dues_range = request.args.get('dues_range', '').strip()
    sort_by = request.args.get('sort_by', '').strip()

    wb = openpyxl.load_workbook(students_file)
    ws = wb.active
    all_students = list(ws.iter_rows(min_row=2, values_only=True))

    exact_match = None
    similar_students = []
    filtered_students = [] # Initialize filtered_students

    if search_query:
        # This temporary list will hold matches from the loop
        # We'll assign it to filtered_students after the loop if it's populated,
        # or handle the exact_match logic.
        current_search_results = []
        for student_tuple in all_students:
            s_id = str(student_tuple[0]).lower() if len(student_tuple) > 0 and student_tuple[0] is not None else ""
            s_name = str(student_tuple[1]).lower() if len(student_tuple) > 1 and student_tuple[1] is not None else ""
            s_father_name = str(student_tuple[2]).lower() if len(student_tuple) > 2 and student_tuple[2] is not None else ""
            # s_contact = str(student_tuple[3]).lower() if len(student_tuple) > 3 and student_tuple[3] is not None else "" # Not used in this part of search

            if search_query.isdigit() or search_query.startswith('s'): # Search by ID
                if search_query == s_id:
                    exact_match = student_tuple # Found exact match by ID
                    # Find similar students (same father's name or contact)
                    if exact_match: # Should always be true here
                        similar_students = [
                            s for s in all_students
                            if len(s) > 3 and s != exact_match and (
                                (s[2] == exact_match[2] and str(s[2]).strip()) or
                                (str(s[3]) == str(exact_match[3]) and str(s[3]).strip())
                            )
                        ]
                    # For exact ID match, we prioritize it and its similar ones
                    # So, we can break the loop and set filtered_students outside
                    break 
                elif search_query in s_id: # Partial ID match
                    current_search_results.append(student_tuple)
            elif (search_query in s_id or 
                  search_query in s_name or 
                  search_query in s_father_name): # General search
                current_search_results.append(student_tuple)
        
        if exact_match: # If an exact ID match was found
            filtered_students = [exact_match] + similar_students
        elif current_search_results: # If no exact ID match, but other matches found
            filtered_students = current_search_results
        elif search_query: # If search query was provided but no matches at all
             flash(f"No student found for query: {search_query}", "info")
             filtered_students = [] # Ensure it's an empty list

    else: # No search query
        filtered_students = all_students

    if grade_filter:
        filtered_students = [s for s in filtered_students if len(s) > 4 and str(s[4]).strip() == grade_filter]

    if dues_range:
        try:
            dues_range_clean = dues_range.replace(" ", "")
            temp_filtered_students = []
            for student_tuple in filtered_students:
                if len(student_tuple) > 8 and student_tuple[8] is not None:
                    try:
                        student_dues = float(student_tuple[8])
                        if dues_range_clean.endswith('+'):
                            min_dues = float(dues_range_clean[:-1])
                            if student_dues >= min_dues:
                                temp_filtered_students.append(student_tuple)
                        else:
                            min_dues, max_dues = map(float, dues_range_clean.split('-'))
                            if min_dues <= student_dues < max_dues:
                                temp_filtered_students.append(student_tuple)
                    except ValueError:
                        # print(f"Warning: Could not parse dues '{student_tuple[8]}' for student ID '{student_tuple[0]}'")
                        continue 
            filtered_students = temp_filtered_students
        except Exception as e:
            flash(f"Invalid dues range format: {e}", "error")


    grade_order = ["PG", "NUR", "PREP", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]
    if sort_by == 'dues':
        def get_dues_key(student_tuple):
            try:
                return float(student_tuple[8] if len(student_tuple) > 8 and student_tuple[8] is not None else 0)
            except ValueError:
                return 0 
        sorted_students = sorted(filtered_students, key=get_dues_key, reverse=True)
    else:
        def get_grade_key(student_tuple):
            grade_val = str(student_tuple[4]) if len(student_tuple) > 4 and student_tuple[4] is not None else ""
            return grade_order.index(grade_val) if grade_val in grade_order else len(grade_order)
        sorted_students = sorted(filtered_students, key=get_grade_key)


    categorized_students = defaultdict(lambda: {'students': [], 'totals': {'fee': 0.0, 'paid_fee': 0.0, 'dues': 0.0}})
    grand_total_metrics = {'fee': 0.0, 'paid_fee': 0.0, 'dues': 0.0}

    for s_data in sorted_students:
        s_grade = str(s_data[4]) if len(s_data) > 4 and s_data[4] else "N/A"
        s_monthly_fee = float(s_data[5]) if len(s_data) > 5 and s_data[5] else 0.0
        s_months_paid_str = s_data[6] if len(s_data) > 6 and s_data[6] else ""
        s_months_paid_count = len([m for m in s_months_paid_str.split(",") if m.strip()]) if s_months_paid_str else 0
        s_paid_fee = float(s_data[7]) if len(s_data) > 7 and s_data[7] else 0.0
        s_dues = float(s_data[8]) if len(s_data) > 8 and s_data[8] else 0.0
        s_annual_charge = float(s_data[10]) if len(s_data) > 10 and s_data[10] else 0.0
        s_paid_annual_charge = float(s_data[11]) if len(s_data) > 11 and s_data[11] else 0.0
        s_admission_fee = float(s_data[12]) if len(s_data) > 12 and s_data[12] else 0.0
        s_paid_admission_fee = float(s_data[13]) if len(s_data) > 13 and s_data[13] else 0.0
        s_exam_charge = float(s_data[15]) if len(s_data) > 15 and s_data[15] else 0.0
        s_paid_exam_charge = float(s_data[16]) if len(s_data) > 16 and s_data[16] else 0.0

        student_total_expected = (s_monthly_fee * s_months_paid_count) + s_annual_charge + s_admission_fee + s_exam_charge
        student_total_paid = s_paid_fee + s_paid_annual_charge + s_paid_admission_fee + s_paid_exam_charge
        
        categorized_students[s_grade]['students'].append(s_data)
        categorized_students[s_grade]['totals']['fee'] += student_total_expected
        categorized_students[s_grade]['totals']['paid_fee'] += student_total_paid
        categorized_students[s_grade]['totals']['dues'] += s_dues 

        grand_total_metrics['fee'] += student_total_expected
        grand_total_metrics['paid_fee'] += student_total_paid
        grand_total_metrics['dues'] += s_dues
        
    return render_template(
        'students.html',
        categorized_students=categorized_students,
        total_fee=grand_total_metrics['fee'],
        total_paid_fee=grand_total_metrics['paid_fee'],
        total_dues=grand_total_metrics['dues'],
        search_query=request.args.get('search', ''),
        grade_filter=grade_filter,
        dues_range=dues_range,
        sort_by=sort_by,
        grade_order=grade_order
    )


@app.route('/students/delete/<string:student_id>', methods=['POST'])
def delete_student(student_id):
    print(f"--- Starting deletion process for student_id: '{student_id}' ---") # Debug
    wb = openpyxl.load_workbook(students_file)
    ws = wb.active
    row_to_delete = None
    picture_filename = None
    student_id_stripped = student_id.strip()
    print(f"Stripped ID to search for: '{student_id_stripped}'") # Debug

    # Find the student row and get picture filename
    for row in ws.iter_rows(min_row=2):
        # row[0] is the cell for ID, row[18] is the cell for Picture
        cell_value = row[0].value
        cell_id_str = str(cell_value).strip() if cell_value is not None else ""
        
        print(f"Checking row {row[0].row}: Found ID '{cell_id_str}'") # Debug

        if cell_id_str == student_id_stripped:
            print(f"MATCH FOUND: Student '{student_id_stripped}' found at Excel row {row[0].row}") # Debug
            row_to_delete = row
            if len(row) > 18 and row[18].value:
                picture_filename = str(row[18].value)
                print(f"Associated picture found: '{picture_filename}'") # Debug
            else:
                print("No associated picture found for this student.") # Debug
            break  # Found the student, no need to search further

    if row_to_delete is None:
        print(f"ERROR: Student with ID '{student_id_stripped}' was not found after checking all rows.") # Debug
        flash("Student not found. Please check the ID.", "error")
        return redirect(url_for('manage_students'))

    # Delete the picture file if it exists
    if picture_filename:
        try:
            picture_path = os.path.join(app.config['UPLOAD_FOLDER'], picture_filename)
            print(f"Attempting to delete picture file at: '{picture_path}'") # Debug
            if os.path.exists(picture_path):
                os.remove(picture_path)
                print("Picture file deleted successfully.") # Debug
            else:
                print("Picture file not found at the specified path.") # Debug
        except OSError as e:
            # Log the error or flash a more specific message if needed
            print(f"ERROR deleting picture file: {e}") # Debug
            flash(f"Error deleting picture file: {e}", "error")

    # Delete the row from the Excel sheet using its row index
    row_index_to_delete = row_to_delete[0].row
    print(f"Attempting to delete Excel row number: {row_index_to_delete}") # Debug
    ws.delete_rows(row_index_to_delete)
    print("Row deleted from worksheet object.") # Debug
    
    try:
        wb.save(students_file)
        print("Excel file saved successfully.") # Debug
    except Exception as e:
        print(f"ERROR saving Excel file: {e}") # Debug
        flash(f"Could not save changes to the data file: {e}", "error")
        return redirect(url_for('manage_students'))

    flash("Student deleted successfully!", "success")
    print(f"--- Finished deletion process for student_id: '{student_id}' ---") # Debug
    return redirect(url_for('manage_students'))


@app.route('/students/modify/<string:student_id>', methods=['GET', 'POST'])
def modify_student(student_id):
    wb = openpyxl.load_workbook(students_file)
    ws = wb.active
    student_row_cells = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        if str(row[0].value).strip() == student_id:
            student_row_cells = row
            break
    if student_row_cells is None:
        flash("Student not found.", "error")
        return redirect(url_for('manage_students'))

    if request.method == 'POST':
        # Existing numeric values (fallback to 0.0)
        ex_monthly_fee = float(student_row_cells[5].value) if student_row_cells[5].value else 0.0
        ex_paid_fee = float(student_row_cells[7].value) if student_row_cells[7].value else 0.0
        ex_annual_charge = float(student_row_cells[10].value) if student_row_cells[10].value else 0.0
        ex_paid_annual_charge = float(student_row_cells[11].value) if student_row_cells[11].value else 0.0
        ex_admission_fee = float(student_row_cells[12].value) if student_row_cells[12].value else 0.0
        ex_paid_admission_fee = float(student_row_cells[13].value) if student_row_cells[13].value else 0.0
        ex_exam_charge = float(student_row_cells[15].value) if student_row_cells[15].value else 0.0
        ex_paid_exam_charge = float(student_row_cells[16].value) if student_row_cells[16].value else 0.0

        # Form inputs (use existing values as fallback where appropriate)
        new_monthly_fee = float(request.form['monthly_fee']) if request.form.get('monthly_fee', '').strip() else ex_monthly_fee
        additional_paid_fee = float(request.form['paid_fee']) if request.form.get('paid_fee', '').strip() else 0.0
        new_annual_charge = float(request.form['annual_charge']) if request.form.get('annual_charge', '').strip() else ex_annual_charge
        new_paid_annual_charge = float(request.form['paid_annual_charge']) if request.form.get('paid_annual_charge', '').strip() else ex_paid_annual_charge
        new_admission_fee = float(request.form['admission_fee']) if request.form.get('admission_fee', '').strip() else ex_admission_fee
        new_paid_admission_fee = float(request.form['paid_admission_fee']) if request.form.get('paid_admission_fee', '').strip() else ex_paid_admission_fee
        new_admission_date = request.form.get('admission_date', '').strip() or (student_row_cells[14].value if student_row_cells[14].value else "")
        new_exam_charge = float(request.form['exam_charge']) if request.form.get('exam_charge', '').strip() else ex_exam_charge
        new_paid_exam_charge = float(request.form['paid_exam_charge']) if request.form.get('paid_exam_charge', '').strip() else ex_paid_exam_charge
        new_payment_date_for_months = request.form.get('paid_date', '').strip()

        # Months handling
        newly_selected_months_paid = request.form.getlist('months_paid') or []
        months_to_delete = request.form.getlist('months_to_delete') or []
        all_possible_months = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
                               'August', 'September', 'October', 'November', 'December']

        # Current stored months and dates
        current_months_paid_str = student_row_cells[6].value or ""
        current_months_paid_list = [m.strip() for m in current_months_paid_str.split(",") if m.strip()]
        current_paid_dates_str = student_row_cells[9].value or ""
        current_paid_dates_list = [d.strip() for d in current_paid_dates_str.split(",") if d.strip()]

        # Build a map month -> existing date (if present). We use indices to align.
        temp_paid_dates_map = {}
        for i, month_name in enumerate(current_months_paid_list):
            if i < len(current_paid_dates_list):
                temp_paid_dates_map[month_name] = current_paid_dates_list[i]
            else:
                temp_paid_dates_map[month_name] = ""

        # Combine months: keep order of first appearance (existing then newly selected)
        combined_months = list(dict.fromkeys(current_months_paid_list + newly_selected_months_paid))
        # Remove months that user requested deletion
        final_months_paid_list = [m for m in combined_months if m not in months_to_delete and m.strip()]
        # Sort according to academic months order (so display is logical)
        final_months_paid_list.sort(key=lambda month: all_possible_months.index(month) if month in all_possible_months else 99)

        # Build final paid dates list:
        # - Preserve existing month->date mappings.
        # - Assign new_payment_date_for_months ONLY to months that are newly added in this request
        #   (i.e., they are in newly_selected_months_paid but NOT in current_months_paid_list).
        final_paid_dates_list = []
        for month_name in final_months_paid_list:
            was_existing_month = month_name in current_months_paid_list
            is_newly_selected_month = month_name in newly_selected_months_paid

            # If it's a newly added month (not existing before) and a new date is provided, use the new date.
            if is_newly_selected_month and not was_existing_month and new_payment_date_for_months:
                final_paid_dates_list.append(new_payment_date_for_months)
            # Otherwise, use the existing date from our map. If not found, use an empty string.
            else:
                final_paid_dates_list.append(temp_paid_dates_map.get(month_name, ""))

        # Update Excel row cells
        student_row_cells[5].value = new_monthly_fee
        student_row_cells[6].value = ", ".join(final_months_paid_list)
        student_row_cells[9].value = ", ".join(final_paid_dates_list)

        # Update paid amount (incremental)
        updated_total_monthly_paid_fee = ex_paid_fee + additional_paid_fee
        student_row_cells[7].value = updated_total_monthly_paid_fee

        # Update other charges and payments
        student_row_cells[10].value = new_annual_charge
        student_row_cells[11].value = new_paid_annual_charge
        student_row_cells[12].value = new_admission_fee
        student_row_cells[13].value = new_paid_admission_fee
        student_row_cells[14].value = new_admission_date
        student_row_cells[15].value = new_exam_charge
        student_row_cells[16].value = new_paid_exam_charge

        # Recalculate dues
        total_expected_monthly_fees = len([m for m in final_months_paid_list if m.strip()]) * new_monthly_fee
        total_expected_overall = total_expected_monthly_fees + new_annual_charge + new_admission_fee + new_exam_charge
        total_paid_overall = updated_total_monthly_paid_fee + new_paid_annual_charge + new_paid_admission_fee + new_paid_exam_charge
        student_row_cells[8].value = total_expected_overall - total_paid_overall

        # Picture handling (optional)
        picture_file = request.files.get('picture')
        if picture_file and picture_file.filename != "":
            if allowed_file(picture_file.filename):
                old_picture_filename = student_row_cells[18].value
                if old_picture_filename:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], str(old_picture_filename)))
                    except OSError:
                        pass
                filename = secure_filename(picture_file.filename)
                new_picture_filename = f"{student_id}_{filename}"
                picture_file.save(os.path.join(app.config['UPLOAD_FOLDER'], new_picture_filename))
                student_row_cells[18].value = new_picture_filename
            else:
                flash("Invalid file type for picture. Allowed types: png, jpg, jpeg, gif", "error")
                return redirect(url_for('modify_student', student_id=student_id))

        wb.save(students_file)
        flash("Student information updated successfully!", "success")
        return redirect(url_for('manage_students'))

    # GET: prepare data for rendering the form
    student_data = {
        "id": student_row_cells[0].value, "name": student_row_cells[1].value,
        "fathers_name": student_row_cells[2].value, "contact_number": student_row_cells[3].value,
        "grade": student_row_cells[4].value, "monthly_fee": student_row_cells[5].value,
        "paid_fee": student_row_cells[7].value, "annual_charge": student_row_cells[10].value,
        "paid_annual_charge": student_row_cells[11].value, "admission_fee": student_row_cells[12].value,
        "paid_admission_fee": student_row_cells[13].value, "admission_date": student_row_cells[14].value,
        "exam_charge": student_row_cells[15].value, "paid_exam_charge": student_row_cells[16].value,
        "months_paid": [m.strip() for m in (student_row_cells[6].value or "").split(",") if m.strip()],
        "paid_dates_str": student_row_cells[9].value, "picture": student_row_cells[18].value
    }
    all_possible_months = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
                           'August', 'September', 'October', 'November', 'December']
    return render_template('test.html', student=student_data, months=all_possible_months)

@app.route('/students/report/<string:student_id>', methods=['GET', 'POST'])
def student_report(student_id):
    student_id_stripped = student_id.strip()
    wb = openpyxl.load_workbook(students_file)
    ws = wb.active
    student_row_cells = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value is not None and str(row[0].value).strip() == student_id_stripped:
            student_row_cells = row
            break
    if student_row_cells is None:
        flash("Student not found. Please check the ID.", "error")
        return redirect(url_for('manage_students'))

    if request.method == 'POST':
        submitted_remarks = request.form.getlist('remarks') 
        student_row_cells[17].value = ", ".join(submitted_remarks) 
        wb.save(students_file)
        flash("Remarks saved successfully!", "success")
        return redirect(url_for('student_report', student_id=student_id_stripped))

    s_data = {
        "id": student_row_cells[0].value, "name": student_row_cells[1].value,
        "father_name": student_row_cells[2].value, "contact_number": student_row_cells[3].value,
        "grade": student_row_cells[4].value,
        "monthly_fee": float(student_row_cells[5].value) if student_row_cells[5].value else 0.0,
        "months_paid_str": student_row_cells[6].value if student_row_cells[6].value else "",
        "paid_fee": float(student_row_cells[7].value) if student_row_cells[7].value else 0.0,
        "dues": float(student_row_cells[8].value) if student_row_cells[8].value else 0.0,
        "paid_dates_str": student_row_cells[9].value if student_row_cells[9].value else "",
        "annual_charge": float(student_row_cells[10].value) if student_row_cells[10].value else 0.0,
        "paid_annual_charge": float(student_row_cells[11].value) if student_row_cells[11].value else 0.0,
        "admission_fee": float(student_row_cells[12].value) if student_row_cells[12].value else 0.0,
        "paid_admission_fee": float(student_row_cells[13].value) if student_row_cells[13].value else 0.0,
        "admission_date": student_row_cells[14].value if student_row_cells[14].value else "",
        "exam_charge": float(student_row_cells[15].value) if student_row_cells[15].value else 0.0,
        "paid_exam_charge": float(student_row_cells[16].value) if student_row_cells[16].value else 0.0,
        "remarks_str": student_row_cells[17].value if student_row_cells[17].value else "",
        "picture": student_row_cells[18].value if len(student_row_cells) > 18 and student_row_cells[18].value else ""
    }
    s_data["months_paid"] = [m.strip() for m in s_data["months_paid_str"].split(',') if m.strip()]
    s_data["paid_dates"] = [d.strip() for d in s_data["paid_dates_str"].split(',') if d.strip()]
    s_data["remarks"] = [r.strip() for r in s_data["remarks_str"].split(',') if r.strip()]

    payment_history = []
    for i, month in enumerate(s_data["months_paid"]):
        payment_entry = {
            "month": month,
            "date_paid": s_data["paid_dates"][i] if i < len(s_data["paid_dates"]) else "N/A",
            "amount": s_data["monthly_fee"], 
            "remark": s_data["remarks"][i] if i < len(s_data["remarks"]) else "" 
        }
        payment_history.append(payment_entry)

    total_expected_fee = (len(s_data["months_paid"]) * s_data["monthly_fee"]) + \
                         s_data["annual_charge"] + s_data["admission_fee"] + s_data["exam_charge"]
    total_actually_paid = s_data["paid_fee"] + s_data["paid_annual_charge"] + \
                          s_data["paid_admission_fee"] + s_data["paid_exam_charge"]

    return render_template(
        'student_report1.html', 
        student=s_data, payment_history=payment_history,
        total_expected_fee=total_expected_fee, 
        total_actually_paid=total_actually_paid, 
        calculated_dues = s_data["dues"] 
    )


grade_order = ["PG", "NUR", "PREP", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]
grade_sort_key = {grade: i for i, grade in enumerate(grade_order)}
academic_months = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
                   'August', 'September', 'October', 'November', 'December']

def parse_date(date_str):
    if not date_str: return None
    try:
        return datetime.strptime(str(date_str).split(" ")[0], '%Y-%m-%d').date() 
    except ValueError:
        try: 
            return datetime.strptime(str(date_str).split(" ")[0], '%d-%m-%Y').date()
        except ValueError:
            return None

@app.route('/fee_collection', methods=['GET'])
def fee_collection():
    # --- 1. Read query parameters ---
    start_date_str   = request.args.get('start_date', '').strip()
    end_date_str     = request.args.get('end_date', '').strip()
    grade_filter     = request.args.get('grade', '').strip()
    month_filter     = request.args.get('month', '').strip()
    dues_range       = request.args.get('dues_range', '').strip()
    view_type        = request.args.get('view_type', 'paid').strip().lower()
    search_q         = request.args.get('search_student', '').lower().strip()

    # --- 2. Parse dates ---
    start_date = parse_date(start_date_str)
    end_date   = parse_date(end_date_str)

    raw_records = []
    all_grades = set()
    grand_total = 0.0

    # --- 3. Load workbook ---
    try:
        wb = openpyxl.load_workbook(students_file, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f"Could not open data file: {e}", 'error')
        return render_template('fee_collection.html', all_records=[], all_grades=[], months=academic_months,
                               grand_total=0, view_type=view_type, error=str(e),
                               start_date=start_date_str, end_date=end_date_str,
                               grade_filter=grade_filter, month_filter=month_filter,
                               dues_range=dues_range, search_student=search_q)

    # --- 4. Build raw per-month entries ---
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid   = str(row[0] or '').strip()
        name  = str(row[1] or '').strip()
        grade = str(row[4] or '').strip()
        if not grade:
            continue
        all_grades.add(grade)
        if grade_filter and grade != grade_filter:
            continue

        fee         = float(row[5] or 0)
        admit_dt    = parse_date(str(row[14] or ''))
        paid_months = [m.strip() for m in str(row[6] or '').split(',') if m.strip()]
        paid_dates  = [d.strip() for d in str(row[9] or '').split(',') if d.strip()]

        if view_type == 'paid':
            for i, m in enumerate(paid_months):
                if month_filter and month_filter.lower() != m.lower():
                    continue
                dt_str = paid_dates[i] if i < len(paid_dates) else ''
                dt_obj = parse_date(dt_str)
                # If user asked date range, exclude entries that have no parsed paid date
                if start_date and (dt_obj is None or dt_obj < start_date):
                    continue
                if end_date and (dt_obj is None or dt_obj > end_date):
                    continue
                rec = {
                    'student_id': sid,
                    'student':    name,
                    'grade':      grade,
                    'parsed_date': dt_obj,
                    'display_date': dt_obj.isoformat() if dt_obj else dt_str,
                    'month':       m,
                    'amount':      fee,
                    'type':        'paid'
                }
                raw_records.append(rec)
                grand_total += fee


        else:
            # build paid set
            paid_set = {}
            for i, m in enumerate(paid_months):
                if i < len(paid_dates):
                    pd = parse_date(paid_dates[i])
                    if pd:
                        paid_set[(pd.year, m.lower())] = True
            # determine years
            current_year = datetime.now().year
            if start_date and end_date:
                years = range(start_date.year, end_date.year + 1)
            elif start_date:
                years = range(start_date.year, current_year + 1)
            elif end_date:
                sy = admit_dt.year if admit_dt else current_year
                years = range(sy, end_date.year + 1)
            else:
                sy = admit_dt.year if admit_dt else current_year
                years = range(sy, current_year + 1)
            # generate unpaid
            for y in years:
                for idx, mn in enumerate(academic_months):
                    m_date = date(y, idx + 1, 1)
                    if admit_dt and m_date < date(admit_dt.year, admit_dt.month, 1):
                        continue
                    if not end_date and m_date > date.today():
                        continue
                    if start_date and m_date < date(start_date.year, start_date.month, 1):
                        continue
                    if end_date and m_date > date(end_date.year, end_date.month, 1):
                        continue
                    if month_filter and month_filter.lower() != mn.lower():
                        continue
                    if not paid_set.get((y, mn.lower()), False):
                        rec = {
                            'student_id': sid,
                            'student':    name,
                            'grade':      grade,
                            'parsed_date': m_date,
                            'display_month': f"{mn} {y}",
                            'amount':       fee,
                            'type':        'unpaid'
                        }
                        raw_records.append(rec)
                        grand_total += fee

    # --- 5. Apply student search filter ---
    if search_q:
        raw_records = [r for r in raw_records
                       if r['student_id'].lower() == search_q
                          or search_q in r['student'].lower()]
        if not raw_records:
            flash(f"No {view_type} records for '{search_q}'", 'info')

    # --- 6. Aggregate unpaid summary ---
    unpaid_summary = (view_type == 'unpaid' and not search_q)
    if unpaid_summary:
        summary = {}
        for r in raw_records:
            sid = r['student_id']
            summary.setdefault(sid, {'student':r['student'],'grade':r['grade'],'months':[],'total_due':0.0})
            summary[sid]['months'].append(r['display_month'])
            summary[sid]['total_due'] += r['amount']
        raw_records = [
            {'student_id':sid,'student':v['student'],'grade':v['grade'],'count':len(v['months']),'months_list':v['months'],'total_due':v['total_due']}
            for sid,v in summary.items()
        ]
        grand_total = sum(v['total_due'] for v in raw_records)
        # filter by dues_range
        if dues_range:
            dr = dues_range.replace(' ', '')
            filtered = []
            try:
                if dr.endswith('+'):
                    min_due = float(dr[:-1])
                    filtered = [r for r in raw_records if r['total_due'] >= min_due]
                else:
                    low, high = map(float, dr.split('-'))
                    filtered = [r for r in raw_records if low <= r['total_due'] < high]
                raw_records = filtered
            except ValueError:
                flash(f"Invalid dues range: {dues_range}", 'error')

    # --- 7. Sort ---
    def sort_key(r):
        dt = r.get('parsed_date')
        if dt is None:
            dt = date.min if r.get('type') == 'paid' else date.max
        return (
            grade_sort_key.get(r.get('grade', ''), len(grade_order)),
            dt,
            r.get('student', '').lower()
        )
    final = sorted(raw_records, key=sort_key)
    sorted_grades = sorted(all_grades, key=lambda g: grade_sort_key.get(g, len(grade_order)))

    # --- 8. Render template ---
    return render_template('fee_collection.html',
                            all_records=final,
                            all_grades=sorted_grades,
                            months=academic_months,
                            start_date=start_date_str,
                            end_date=end_date_str,
                            grade_filter=grade_filter,
                            month_filter=month_filter,
                            dues_range=dues_range,
                            view_type=view_type,
                            grand_total=grand_total,
                            search_student=search_q,
                            unpaid_summary=unpaid_summary)

# if __name__ == '__main__':
#     # For local development without webview
#     # threading.Thread(target=start_flask, daemon=True).start()
#     # webview.create_window("Student Fee Management", "http://127.0.0.1:5000/", width=1000, height=700)
#     # webview.start()
#     app.run(debug=True) # Standard Flask debug mode

  # For secure file names
app.secret_key = "secret_key"  
data_dir = "data"

if not os.path.exists(data_dir):
    os.makedirs(data_dir)

# Define upload folder for pictures (stored in static/uploads)
UPLOAD_FOLDER = os.path.join(app.root_path, 'static', 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Allowed file extensions for student pictures.
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Paths for Excel files
comp_file = os.path.join(data_dir, "comp.xlsx")

# Initialize Excel file with updated headers including "Admission Date" and "Picture"
def initialize_excel_file(file_path, headers):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)  # Add header row
        wb.save(file_path)

# Updated header for computer academy: Removed "Grade", "Annual Charge", "Exam Charge"
# Kept "Admission Date" and "Picture".
initialize_excel_file(comp_file, [
    "ID", "Name", "Father Name", "Contact Number", "Course", "Monthly Fee", "Months Paid",
    "Paid Fee", "Dues", "Paid Dates", "Admission Fee", "Paid Admission Fee", "Admission Date", "Remarks", "Picture"
])





def start_flask():
    app.run(host='127.0.0.1', port=5000)


@app.route('/add_comp', methods=['GET', 'POST'])
def add_comp():
    if request.method == 'POST':
        # Get required fields; ID and Name cannot be empty.
        student_id = request.form.get('id', '').strip()
        name = request.form.get('name', '').strip()
        if not student_id or not name:
            flash("ID and Name are required.", "error")
            return redirect(url_for('add_student'))

        # Get optional fields (if empty, use empty string or 0.0 for numbers)
        fathers_name = request.form.get('fathers_name', '').strip()
        contact_number = request.form.get('contact_number', '').strip()
        course = request.form.get('course', '').strip()
        monthly_fee = float(request.form['monthly_fee']) if request.form.get('monthly_fee', '').strip() else 0.0
        admission_fee = float(request.form['admission_fee']) if request.form.get('admission_fee', '').strip() else 0.0
        paid_admission_fee = float(request.form['paid_admission_fee']) if request.form.get('paid_admission_fee', '').strip() else 0.0
        paid_fee = float(request.form['paid_fee']) if request.form.get('paid_fee', '').strip() else 0.0
        paid_date = request.form.get('paid_date', '').strip()
        admission_date = request.form.get('admission_date', '').strip()  # New field

        # List of months paid
        months_paid = request.form.getlist('months_paid')
        months_paid_str = ", ".join(months_paid)
        # Multiply the monthly fee by the number of months paid.
        total_months_fee = len(months_paid) * monthly_fee

        # Calculate dues using the formula:
        # (Monthly Fee * number of months paid) + Admission Fee
        # minus the total paid amounts.
        dues = (total_months_fee + admission_fee) - (paid_fee + paid_admission_fee)

        # Process picture file upload (optional)
        picture_file = request.files.get('picture')
        picture_filename = ""
        if picture_file and picture_file.filename != "":
            if allowed_file(picture_file.filename):
                filename = secure_filename(picture_file.filename)
                picture_filename = f"{student_id}_{filename}"
                picture_file.save(os.path.join(app.config['UPLOAD_FOLDER'], picture_filename))
            else:
                flash("Invalid file type for picture. Allowed types: png, jpg, jpeg, gif", "error")
                return redirect(url_for('add_student'))

        # Save data to Excel (note the new order of fields; the new "Picture" field is last)
        wb = openpyxl.load_workbook(comp_file)
        ws = wb.active
        ws.append([
            student_id, name, fathers_name, contact_number, course, monthly_fee, months_paid_str,
            paid_fee, dues, paid_date, admission_fee, paid_admission_fee, admission_date, "", picture_filename
        ])
        wb.save(comp_file)

        flash("Student added successfully!", "success")
        return redirect(url_for('comp'))

    return render_template('add_comp.html')


@app.route('/comp', methods=['GET', 'POST'])
def comp():
    # Get query parameters (only available on GET requests)
    search_query = request.args.get('search', '').lower().strip()
    course_filter = request.args.get('course', '').strip()
    dues_range = request.args.get('dues_range', '').strip()  # e.g., "100-500" or "500+"
    sort_by = request.args.get('sort_by', '').strip()  # e.g., "dues"

    if request.method == 'POST':
        # Process form submission to add a new student
        student_id = request.form.get('id', '').strip()
        name = request.form.get('name', '').strip()
        if not student_id or not name:
            flash("ID and Name are required.", "error")
            return redirect(url_for('comp'))

        fathers_name = request.form.get('fathers_name', '').strip()
        contact_number = request.form.get('contact_number', '').strip()
        course = request.form.get('course', '').strip()
        monthly_fee = float(request.form['monthly_fee']) if request.form.get('monthly_fee', '').strip() else 0.0
        admission_fee = float(request.form['admission_fee']) if request.form.get('admission_fee', '').strip() else 0.0
        paid_admission_fee = float(request.form['paid_admission_fee']) if request.form.get('paid_admission_fee', '').strip() else 0.0
        paid_fee = float(request.form['paid_fee']) if request.form.get('paid_fee', '').strip() else 0.0
        paid_date = request.form.get('paid_date', '').strip()
        admission_date = request.form.get('admission_date', '').strip()  # New field

        months_paid = request.form.getlist('months_paid')
        months_paid_str = ", ".join(months_paid)
        total_months_fee = len(months_paid) * monthly_fee

        dues = (total_months_fee + admission_fee) - (paid_fee + paid_admission_fee)

        wb = openpyxl.load_workbook(comp_file)
        ws = wb.active
        ws.append([
            student_id, name, fathers_name, contact_number, course, monthly_fee, months_paid_str,
            paid_fee, dues, paid_date, admission_fee, paid_admission_fee, admission_date, "", ""
        ])
        wb.save(comp_file)

        flash("Student added successfully!", "success")
        return redirect(url_for('comp'))

    # Load student data from Excel (for GET requests)
    wb = openpyxl.load_workbook(comp_file)
    ws = wb.active
    all_students = list(ws.iter_rows(min_row=2, values_only=True))

    # --- SEARCH FILTERING ---
    # Look for the search term in Student ID, Name, or Father's Name
    if search_query:
        filtered_students = [
            student for student in all_students
            if search_query in str(student[0]).lower() or
               search_query in str(student[1]).lower() or
               search_query in str(student[2]).lower()
        ]
    else:
        filtered_students = all_students

    # --- COURSE FILTERING ---
    if course_filter:
        filtered_students = [student for student in filtered_students if str(student[4]).strip() == course_filter]

    # --- DUES RANGE FILTERING ---
    if dues_range:
        try:
            dues_range_clean = dues_range.replace(" ", "")
            if dues_range_clean.endswith('+'):
                min_dues = float(dues_range_clean[:-1])
                filtered_students = [student for student in filtered_students if float(student[8]) >= min_dues]
            else:
                min_dues, max_dues = map(float, dues_range_clean.split('-'))
                filtered_students = [student for student in filtered_students if min_dues <= float(student[8]) < max_dues]
        except Exception as e:
            flash("Invalid dues range format.", "error")

    # --- SORTING ---
    if sort_by == 'dues':
        sorted_students = sorted(filtered_students, key=lambda student: float(student[8]), reverse=True)
    else:
        sorted_students = filtered_students  # No grade to sort by

    # --- CATEGORIZING BY COURSE & Totals Calculation ---
    # Indexes: 5: Monthly Fee, 6: Months Paid, 7: Paid Fee, 8: Dues, 9: Paid Dates,
    # 10: Admission Fee, 11: Paid Admission Fee, 12: Admission Date, 13: Remarks, 14: Picture.
    categorized_students = defaultdict(lambda: {'students': [], 'totals': {'fee': 0.0, 'paid_fee': 0.0, 'dues': 0.0}})
    for student in sorted_students:
        course = str(student[4])
        categorized_students[course]['students'].append(student)

        monthly_fee = student[5] if student[5] else 0.0
        months_paid_count = len((student[6] or "").split(", ")) if student[6] else 0
        admission_fee = student[10] if student[10] else 0.0
        paid_fee = student[7] if student[7] else 0.0
        paid_admission_fee = student[11] if student[11] else 0.0
        dues = student[8] if student[8] else 0.0

        categorized_students[course]['totals']['fee'] += (monthly_fee * months_paid_count) + admission_fee
        categorized_students[course]['totals']['paid_fee'] += paid_fee + paid_admission_fee
        categorized_students[course]['totals']['dues'] += dues

    total_fee = sum(
        (student[5] if student[5] else 0.0) * len((student[6] or "").split(", ")) +
        (student[10] if student[10] else 0.0)
        for student in sorted_students
    )
    total_paid_fee = sum(
        (student[7] if student[7] else 0.0) +
        (student[11] if student[11] else 0.0)
        for student in sorted_students
    )
    total_dues = sum(student[8] if student[8] else 0.0 for student in sorted_students)

    return render_template(
        'comp.html',
        categorized_students=categorized_students,
        total_fee=total_fee,
        total_paid_fee=total_paid_fee,
        total_dues=total_dues,
        search_query=request.args.get('search', ''),
        course_filter=course_filter,
        dues_range=dues_range
    )



@app.route('/comp/delete/<string:student_id>', methods=['POST'])
def delete_comp(student_id):
    wb = openpyxl.load_workbook(comp_file)
    ws = wb.active

    student_row_index = None
    student_id_stripped = student_id.strip() # Ensure no leading/trailing spaces
    picture_filename = None

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if str(row[0]).strip() == student_id_stripped:
            student_row_index = idx
            # Picture is in the 15th column (index 14)
            if len(row) > 14 and row[14]:
                picture_filename = str(row[14])
            break

    if student_row_index is None:
        flash("Student not found. Please check the ID.", "error")
        return redirect(url_for('comp'))

    # Delete picture file if it exists
    if picture_filename:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], picture_filename))
        except OSError as e:
            flash(f"Error deleting picture file {picture_filename}: {e}", "error")

    ws.delete_rows(student_row_index)
    wb.save(comp_file)

    flash("Student deleted successfully!", "success")
    return redirect(url_for('comp'))


@app.route('/comp/modify/<string:student_id>', methods=['GET', 'POST'])
def modify_comp(student_id):
    wb = openpyxl.load_workbook(students_file)
    ws = wb.active

    # Locate the student's row
    student_row = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        if str(row[0].value).strip() == student_id:
            student_row = row
            break
    else:
        flash("Student not found.", "error")
        return redirect(url_for('comp'))

    if request.method == 'POST':
        # Retrieve updated values from the form.
        # Make sure your form includes an input field named "monthly_fee"
        new_monthly_fee = (
            float(request.form['monthly_fee'])
            if request.form.get('monthly_fee', '').strip() else (student_row[5].value or 0.0)
        )
        new_paid_fee = float(request.form['paid_fee']) if request.form.get('paid_fee', '').strip() else 0.0
        new_admission_fee = float(request.form['admission_fee']) if request.form.get('admission_fee', '').strip() else (student_row[10].value or 0.0)
        paid_admission_fee = float(request.form['paid_admission_fee']) if request.form.get('paid_admission_fee', '').strip() else (student_row[11].value or 0.0)
        new_admission_date = request.form.get('admission_date', '').strip()  # New field
        new_paid_date = request.form.get('paid_date', '').strip()

        # Process the months paid.
        new_months_paid = request.form.getlist('months_paid')
        months_to_delete = request.form.getlist('months_to_delete')
        months = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
                  'August', 'September', 'October', 'November', 'December']
        # Split the stored months string and combine with new ones.
        current_months = (student_row[6].value or "").split(", ")
        current_months = [m for m in current_months if m.strip()]
        combined_months = current_months + new_months_paid

        # Remove duplicates by using set, sort them according to our months order.
        updated_months_paid = sorted(set(combined_months), key=lambda month: months.index(month) if month in months else 100)
        # Remove any months that are marked for deletion.
        updated_months_paid = [month for month in updated_months_paid if month not in months_to_delete]
        updated_months_paid = [month for month in updated_months_paid if month.strip()]

        # Update paid dates.
        current_paid_dates = (student_row[9].value or "").split(", ")
        current_paid_dates = [d for d in current_paid_dates if d.strip()]
        updated_paid_dates = current_paid_dates + [new_paid_date] * len(new_months_paid)
        aligned_paid_dates = updated_paid_dates[:len(updated_months_paid)]

        # Update the Excel row.
        # Column indexes based on our header:
        # 0: ID, 1: Name, 2: Father Name, 3: Contact Number, 4: Course,
        # 5: Monthly Fee, 6: Months Paid, 7: Paid Fee, 8: Dues, 9: Paid Dates,
        # 10: Admission Fee, 11: Paid Admission Fee, 12: Admission Date, 13: Remarks, 14: Picture.
        student_row[5].value = new_monthly_fee                # Update monthly fee.
        student_row[6].value = ", ".join(updated_months_paid)
        student_row[9].value = ", ".join(aligned_paid_dates)

        total_months_fee = len(updated_months_paid) * new_monthly_fee

        student_row[7].value = new_paid_fee
        student_row[10].value = new_admission_fee
        student_row[11].value = paid_admission_fee
        student_row[12].value = new_admission_date  # Admission Date.

        # Recalculate dues:
        student_row[8].value = (total_months_fee + new_admission_fee) - (
            new_paid_fee + paid_admission_fee
        )

        # Process picture file upload (optional update)
        picture_file = request.files.get('picture')
        if picture_file and picture_file.filename != "":
            if allowed_file(picture_file.filename):
                filename = secure_filename(picture_file.filename)
                picture_filename = f"{student_id}_{filename}"
                picture_file.save(os.path.join(app.config['UPLOAD_FOLDER'], picture_filename))
                student_row[14].value = picture_filename
            else:
                flash("Invalid file type for picture. Allowed types: png, jpg, jpeg, gif", "error")
                return redirect(url_for('modify_comp', student_id=student_id))

        wb.save(students_file)
        flash("Student information updated successfully!", "success")
        return redirect(url_for('comp'))

    # For GET, prepare the student data to populate the form.
    student_data = {
        "id": student_row[0].value,
        "name": student_row[1].value,
        "fathers_name": student_row[2].value,
        "contact_number": student_row[3].value,
        "course": student_row[4].value,
        "monthly_fee": student_row[5].value,
        "admission_fee": student_row[10].value,
        "paid_fee": student_row[7].value,
        "paid_admission_fee": student_row[11].value,
        "admission_date": student_row[12].value,  # New field
        "months_paid": [m for m in (student_row[6].value or "").split(", ") if m.strip()],
        "paid_date": student_row[9].value,
        "picture": student_row[14].value  # Include picture filename (if any)
    }

    months = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
              'August', 'September', 'October', 'November', 'December']

    return render_template('comp_manage.html', student=student_data, months=months)


@app.route('/comp/report/<string:student_id>', methods=['GET', 'POST'])
def comp_report(student_id):
    student_id = student_id.strip()
    wb = openpyxl.load_workbook(comp_file)
    ws = wb.active

    student_row = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value is not None and str(row[0].value).strip() == student_id:
            student_row = row
            break

    if student_row is None:
        flash("Student not found. Please check the ID.", "error")
        return redirect(url_for('comp'))

    if request.method == 'POST':
        submitted_remarks = request.form.getlist('remarks')
        existing_remarks = student_row[13].value.split(", ") if student_row[13].value else []
        updated_remarks = [
            remark if i < len(submitted_remarks) else existing_remarks[i]
            for i, remark in enumerate(submitted_remarks)
        ]
        student_row[13].value = ", ".join(updated_remarks)
        wb.save(comp_file)
        flash("Remarks saved successfully!", "success")
        return redirect(url_for('comp_report', student_id=student_id))

    student_data = {
        "id": student_row[0].value,
        "name": student_row[1].value,
        "father_name": student_row[2].value,
        "contact_number": student_row[3].value,
        "course": student_row[4].value,
        "monthly_fee": float(student_row[5].value) if student_row[5].value else 0.0,
        "months_paid": student_row[6].value.split(", ") if student_row[6].value else [],
        "paid_fee": float(student_row[7].value) if student_row[7].value else 0.0,
        "dues": float(student_row[8].value) if student_row[8].value else 0.0,
        "paid_dates": student_row[9].value.split(", ") if student_row[9].value else [],
        "admission_fee": float(student_row[10].value) if student_row[10].value else 0.0,
        "paid_admission_fee": float(student_row[11].value) if student_row[11].value else 0.0,
        "admission_date": student_row[12].value,  # New field
        "remarks": student_row[13].value.split(", ") if student_row[13].value else [],
        "picture": student_row[14].value if len(student_row) > 14 else ""

    }

    payment_history = []
    for i, month in enumerate(student_data["months_paid"]):
        payment_entry = {
            "month": month,
            "date_paid": student_data["paid_dates"][i] if i < len(student_data["paid_dates"]) else "N/A",
            "remark": student_data["remarks"][i] if i < len(student_data["remarks"]) else ""
        }
        payment_history.append(payment_entry)

    total_expected_fee = (len(student_data["months_paid"]) * student_data["monthly_fee"]) + \
                           student_data["admission_fee"]

    return render_template(
        'comp_report.html',
        student=student_data,
        payment_history=payment_history,
        total_expected_fee=total_expected_fee
    )


@app.route('/comp_fee', methods=['GET'])
def comp_fee():
    """
    Displays fee collection data date-wise.
    """
    filter_date = request.args.get('filter_date', '').strip()

    fee_collection_summary = defaultdict(lambda: {'count': 0, 'total_fee': 0.0, 'students': []})

    wb = openpyxl.load_workbook(students_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        monthly_fee = float(row[5]) if row[5] else 0.0
        paid_dates_str = row[9] if row[9] else ''
        paid_dates = [date.strip() for date in paid_dates_str.split(",") if date.strip()]

        for date in paid_dates:
            if filter_date and date != filter_date:
                continue
            fee_collection_summary[date]['count'] += 1
            fee_collection_summary[date]['total_fee'] += monthly_fee
            fee_collection_summary[date]['students'].append(row[1])

    summary_list = sorted(fee_collection_summary.items(), key=lambda x: x[0])

    return render_template(
        'comp_fee.html',
        summary_list=summary_list,
        filter_date=filter_date,
    )

# Function to start the Flask server
def start_flask():
    app.run()

if __name__ == '__main__':
    # Start Flask in a separate thread
    flask_thread = threading.Thread(target=start_flask)
    flask_thread.daemon = True
    flask_thread.start()

    # Create a webview window
    webview.create_window('School Software', 'http://127.0.0.1:5000')
    webview.start()