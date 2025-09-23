import os
import openpyxl
import threading
from flask import Flask, render_template, request, redirect, url_for
import webview

app = Flask(__name__)

# Set directory and database file
data_dir = "data"
if not os.path.exists(data_dir):
    os.makedirs(data_dir)

DATABASE = os.path.join(data_dir, "school_data.xlsx")

# Initialize Excel file if it does not exist
if not os.path.exists(DATABASE):
    workbook = openpyxl.Workbook()
    # Teachers sheet with extra fields: Qualification and Contact Number.
    teachers_sheet = workbook.active
    teachers_sheet.title = "Teachers"
    teachers_sheet.append([
        "Teacher Name", "Qualification", "Contact Number", "Joining Date", "Monthly Salary",
        "January Deduction", "January Leaves", "January Paid", 
        "February Deduction", "February Leaves", "February Paid", 
        "March Deduction", "March Leaves", "March Paid", 
        "April Deduction", "April Leaves", "April Paid", 
        "May Deduction", "May Leaves", "May Paid", 
        "June Deduction", "June Leaves", "June Paid", 
        "July Deduction", "July Leaves", "July Paid", 
        "August Deduction", "August Leaves", "August Paid", 
        "September Deduction", "September Leaves", "September Paid", 
        "October Deduction", "October Leaves", "October Paid", 
        "November Deduction", "November Leaves", "November Paid", 
        "December Deduction", "December Leaves", "December Paid"
    ])

    # Report sheet now also stores the teacher's qualification and contact number.
    report_sheet = workbook.create_sheet("Report")
    report_sheet.append([
        "Teacher Name", "Qualification", "Contact Number", "Month", "Deductions", "Leaves", "Paid Salary"
    ])
    workbook.save(DATABASE)

# Helper function for month column mapping with updated indices.
# With the new teacher columns, the first 5 columns are reserved for:
# 0: Teacher Name, 1: Qualification, 2: Contact Number, 3: Joining Date, 4: Monthly Salary.
# So monthly details now start at index 5.
def get_month_columns():
    month_columns = {
        "January":   (5, 6, 7),
        "February":  (8, 9, 10),
        "March":     (11, 12, 13),
        "April":     (14, 15, 16),
        "May":       (17, 18, 19),
        "June":      (20, 21, 22),
        "July":      (23, 24, 25),
        "August":    (26, 27, 28),
        "September": (29, 30, 31),
        "October":   (32, 33, 34),
        "November":  (35, 36, 37),
        "December":  (38, 39, 40)
    }
    return month_columns

# Route to display all teachers and their yearly summary
@app.route('/teachers')
def show_teachers():
    workbook = openpyxl.load_workbook(DATABASE)
    sheet = workbook["Teachers"]
    teachers = []
    month_columns = get_month_columns()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        teacher = {
            "name": row[0],
            "qualification": row[1],
            "contact": row[2],
            "joining_date": row[3],
            "monthly_salary": row[4],
            # For deductions and paid salary, we sum the values from the monthly details.
            "deductions": sum(row[i] or 0 for i in range(5, 41, 3)),
            "paid_salary": sum(row[i+2] or 0 for i in range(5, 41, 3)),
            "total_salary": (row[4] or 0) * 12,
            "months": []
        }
        
        # Add month-wise deduction, leaves, and paid salary details.
        for month, (deduction_col, leaves_col, paid_col) in month_columns.items():
            teacher["months"].append({
                "month": month,
                "deduction": row[deduction_col] or 0,
                "leaves": row[leaves_col] or 0,
                "paid": row[paid_col] or 0
            })

        teachers.append(teacher)

    return render_template("teachers.html", teachers=teachers, months=list(month_columns.keys()))

# Route to add a new teacher (now including qualification and contact number)
@app.route('/teachers/add', methods=['GET', 'POST'])
def add_teacher():
    if request.method == 'POST':
        name = request.form['name']
        qualification = request.form['qualification']
        contact = request.form['contact']
        joining_date = request.form['joining_date']
        monthly_salary = float(request.form['monthly_salary'])

        workbook = openpyxl.load_workbook(DATABASE)
        sheet = workbook["Teachers"]
        # Append a row with 5 teacher details followed by 36 zeros for monthly data.
        row = [name, qualification, contact, joining_date, monthly_salary] + [0] * 36
        sheet.append(row)
        workbook.save(DATABASE)

        return redirect(url_for('show_teachers'))

    return render_template("add_teacher.html")

# Route to update salary and deductions for a specific month
@app.route('/teachers/update', methods=['POST'])
def update_teacher():
    name = request.form['name']
    month = request.form['month']
    attendance_days = int(request.form['attendance_days'])  # Total days teacher was present

    # Define total days per month.
    month_days = {
        "January": 31, "February": 28, "March": 31, "April": 30,
        "May": 31, "June": 30, "July": 31, "August": 31,
        "September": 30, "October": 31, "November": 30, "December": 31
    }

    total_days = month_days[month]
    leaves = total_days - attendance_days  # Auto calculate leaves

    workbook = openpyxl.load_workbook(DATABASE)
    sheet = workbook["Teachers"]
    month_columns = get_month_columns()

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == name:
            deduction_col, leaves_col, paid_col = month_columns[month]
            monthly_salary = row[4].value or 0  # Monthly salary is now at index 4

            # Deduct 4% of the monthly salary per leave.
            deduction = (monthly_salary * 0.04) * leaves
            paid_salary = max(0, monthly_salary - deduction)  # Ensure salary doesn't go negative

            # Update the monthly columns in Excel.
            row[deduction_col].value = deduction
            row[leaves_col].value = leaves
            row[paid_col].value = paid_salary

            # Retrieve the teacher’s qualification and contact number.
            qualification = row[1].value
            contact = row[2].value

            update_report_sheet(name, qualification, contact, month, deduction, leaves, paid_salary)
            break

    workbook.save(DATABASE)
    return redirect(url_for('show_teachers'))

# Helper function to record updates in the Report sheet.
def update_report_sheet(name, qualification, contact, month, deductions, leaves, paid_salary):
    workbook = openpyxl.load_workbook(DATABASE)
    sheet = workbook["Report"]
    sheet.append([name, qualification, contact, month, deductions, leaves, paid_salary])
    workbook.save(DATABASE)

# Route to delete a teacher
@app.route('/teachers/delete/<teacher_name>', methods=['POST'])
def delete_teacher(teacher_name):
    workbook = openpyxl.load_workbook(DATABASE)
    sheet = workbook["Teachers"]

    # Find the teacher by name and delete their row.
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Start at row 2 to skip headers.
        if row[0].value == teacher_name:
            sheet.delete_rows(row_idx)
            break

    workbook.save(DATABASE)
    return redirect(url_for('show_teachers'))

# Route to view and update fees for the next month.
@app.route('/teachers/next_month', methods=['POST'])
def next_month_update():
    name = request.form['name']
    next_month = request.form['next_month']

    workbook = openpyxl.load_workbook(DATABASE)
    sheet = workbook["Teachers"]
    month_columns = get_month_columns()

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == name:
            deduction_col, leaves_col, paid_col = month_columns[next_month]

            # Initialize next month's salary with no deductions or leaves.
            monthly_salary = row[4].value or 0
            row[deduction_col].value = 0  # No deductions initially.
            row[leaves_col].value = 0     # No leaves initially.
            row[paid_col].value = monthly_salary

            qualification = row[1].value
            contact = row[2].value
            update_report_sheet(name, qualification, contact, next_month, 0, 0, monthly_salary)
            break

    workbook.save(DATABASE)
    return redirect(url_for('show_teachers'))

# Route to generate a report for a teacher (now including qualification and contact details)
@app.route('/teachers/report/<teacher_name>', methods=['GET'])
def teacher_report(teacher_name):
    workbook = openpyxl.load_workbook(DATABASE)
    sheet = workbook["Teachers"]
    report = []
    qualification = None
    contact = None
    joining_date = None
    monthly_salary = 0

    month_columns = get_month_columns()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == teacher_name:
            qualification = row[1]
            contact = row[2]
            joining_date = row[3]
            monthly_salary = row[4] if row[4] else 0
            print(f"Found teacher: {teacher_name}, Qualification: {qualification}, Contact: {contact}, Joining Date: {joining_date}, Salary: {monthly_salary}")

            # Loop through each month and add the details.
            for month, (deduction_col, leaves_col, paid_col) in month_columns.items():
                deduction = row[deduction_col] if row[deduction_col] is not None else 0
                leaves = row[leaves_col] if row[leaves_col] is not None else 0
                paid_salary = row[paid_col] if row[paid_col] is not None else 0

                print(f"Processing {month}: Deduction: {deduction}, Leaves: {leaves}, Paid Salary: {paid_salary}")

                report.append({
                    "month": month,
                    "deduction": deduction,
                    "leaves": leaves,
                    "paid_salary": paid_salary
                })
            break

    return render_template("teacher_report.html", teacher_name=teacher_name,
                           qualification=qualification, contact=contact,
                           joining_date=joining_date, monthly_salary=monthly_salary,
                           report=report)

def start_flask():
    app.run()

if __name__ == '__main__':
    # Start Flask in a separate thread.
    flask_thread = threading.Thread(target=start_flask)
    flask_thread.daemon = True
    flask_thread.start()

    # Create a webview window.
    webview.create_window('School Software', 'http://127.0.0.1:5000')
    webview.start()
